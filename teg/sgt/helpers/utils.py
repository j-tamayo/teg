# -*- coding: utf-8 -*-
from sgt.models import *
from openpyxl import load_workbook
from datetime import datetime

def cargar_centros_desde_xls(file):
	"""Función para cargar los centros de inspección desde un archivo xls"""
	try:
		reg = {}
		wb = load_workbook(file)
		ws = wb['centros']
		print "HEEEY",ws.get_highest_row()
		for i in range(2,ws.get_highest_row()+1):
			
			reg['nombre'] = ws['A'+str(i)].value
			reg['estado'] = ws['B'+str(i)].value
			reg['municipio'] = ws['C'+str(i)].value
			reg['direccion'] = ws['D'+str(i)].value
			reg['telefonos'] = ws['E'+str(i)].value
			reg['hora_apertura_manana'] = ws['F'+str(i)].value
			reg['hora_cierre_manana'] = ws['G'+str(i)].value
			reg['hora_apertura_tarde'] = ws['H'+str(i)].value
			reg['hora_cierre_tarde'] = ws['I'+str(i)].value
			
			municipio = Municipio.objects.filter(nombre__icontains = reg['municipio'], estado__nombre__icontains = reg['estado']).first()
			existe_centro = CentroInspeccion.objects.filter(nombre = reg['nombre'], municipio = municipio).count()
			if existe_centro < 1:
				centro = CentroInspeccion(
					nombre = reg['nombre'],
					municipio = municipio,
					direccion = reg['direccion'],
					telefonos = reg['telefonos'],
					hora_apertura_manana = reg['hora_apertura_manana'],
					hora_cierre_manana = reg['hora_cierre_manana'],
					hora_apertura_tarde = reg['hora_apertura_tarde'],
					hora_cierre_tarde = reg['hora_cierre_tarde']
				)
				centro.save()
			else:
				print "Ya se agregó este centro"

	except Exception,e:
		print e
		return False

	return True


def cargar_polizas_desde_xls(file):
	"""Función para cargar los centros de inspección desde un archivo xls"""
	try:
		reg = {}
		wb = load_workbook(file)
		ws = wb['polizas']
		print "HEEEY",ws.get_highest_row()
		for i in range(2,ws.get_highest_row()+1):
			
			reg['numero'] = ws['A'+str(i)].value
			reg['cedula'] = ws['B'+str(i)].value
			reg['fecha_inicio_vigencia'] = ws['C'+str(i)].value
			reg['fecha_fin_vigencia'] = ws['D'+str(i)].value
			reg['fecha_inicio_vigencia'] = datetime.strptime(reg['fecha_inicio_vigencia'], '%d-%m-%Y').date()
			reg['fecha_fin_vigencia'] = datetime.strptime(reg['fecha_fin_vigencia'], '%d-%m-%Y').date()
			print type(reg['fecha_fin_vigencia']),reg['fecha_fin_vigencia']
			usuario = SgtUsuario.objects.filter(cedula = reg['cedula']).first()
			existe_poliza = Poliza.objects.filter(numero = reg['numero']).first()
			if not existe_poliza:
				poliza = Poliza(
					numero = reg['numero'],
					cedula_cliente = reg['cedula'],
					fecha_inicio_vigencia = reg['fecha_inicio_vigencia'],
					fecha_fin_vigencia = reg['fecha_fin_vigencia'],
					usuario = usuario
				)
				poliza.save()

			else:
				existe_poliza.cedula = reg['cedula']
				existe_poliza.fecha_inicio_vigencia = reg['fecha_inicio_vigencia']
				existe_poliza.fecha_fin_vigencia = reg['fecha_fin_vigencia']
				existe_poliza.usuario = usuario
				existe_poliza.save()

	except Exception,e:
		print e
		return False

	return True
		