# -*- coding: utf-8 -*-
from django.db import models
from cuentas.models import SgtUsuario, RolSgt
from datetime import datetime
from django.db.models import Q
import operator
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors

USER_MODEL = SgtUsuario

# Create your models here.
class Estado(models.Model):
	nombre = models.CharField(max_length=255)

	def __unicode__(self):
		return u'%s' % self.nombre


class Municipio(models.Model):
	nombre = models.CharField(max_length=255)
	estado = models.ForeignKey(Estado)

	def __unicode__(self):
		return u'%s' % self.nombre


class CentroInspeccion(models.Model):
	# codigo = models.CharField(max_length=20)
	nombre = models.CharField(max_length=255)
	direccion = models.TextField()
	capacidad = models.IntegerField(default=0)
	telefonos = models.CharField(max_length=255)
	tiempo_atencion = models.IntegerField(default=0)
	municipio = models.ForeignKey(Municipio)
	peritos = models.ManyToManyField('Perito')
	# numero_orden = models.ManyToManyField('NumeroOrden', through='ColaAtencion')
	hora_apertura_manana = models.TimeField(blank=True, null=True)
	hora_cierre_manana = models.TimeField(blank=True, null=True)
	hora_apertura_tarde = models.TimeField(blank=True, null=True)
	hora_cierre_tarde = models.TimeField(blank=True, null=True)
	fechas_no_laborables = models.ManyToManyField('FechaNoLaborable')

	def __unicode__(self):
		return u'%s' % self.nombre


class FechaNoLaborable(models.Model):
	fecha = models.DateField()


class Perito(models.Model):
	apellidos = models.CharField(max_length=200)
	cedula = models.CharField(max_length=100)
	fecha_ingreso = models.DateField()
	nombres = models.CharField(max_length=200)
	sexo = models.IntegerField()
	activo = models.BooleanField(default=True)

	def __unicode__(self):
		return u'%s %s' % (self.nombres, self.apellidos)


class TipoInspeccion(models.Model):
	codigo = models.CharField(max_length=50)
	descripcion = models.TextField(blank=True,null=True)
	nombre = models.CharField(max_length=255)

	def __unicode__(self):
		return u'%s' % self.nombre


class SolicitudInspeccion(models.Model):
	centro_inspeccion = models.ForeignKey(CentroInspeccion)
	fecha_creacion = models.DateTimeField(auto_now_add=True)
	fecha_culminacion = models.DateTimeField(blank=True, null=True)
	perito = models.ForeignKey(Perito, blank=True, null=True)
	tipo_inspeccion = models.ForeignKey(TipoInspeccion)
	estatus = models.ForeignKey('Estatus')
	usuario = models.ForeignKey(USER_MODEL)
	borrada = models.BooleanField(default=False)

	def __unicode__(self):
		return u'%s' % self.tipo_inspeccion.nombre

	def toDict(self):
		return {
			'centro_inspeccion': self.centro_inspeccion.pk,
			'fecha_creacion': datetime.strftime(self.fecha_creacion, '%d-%m-%Y'),
			'perito': (self.perito.nombres + ' ' + self.perito.apellidos) if self.perito is not None else None,
			'tipo_inspeccion': self.tipo_inspeccion.nombre,
			'estatus': self.estatus.nombre,
			'usuario': self.usuario.nombres + ' ' + self.usuario.apellidos,
		}


class NumeroOrden(models.Model):
	asistencia = models.IntegerField(default=0)
	solicitud_inspeccion = models.ForeignKey(SolicitudInspeccion)
	codigo = models.CharField(max_length=50, blank = True, null = True)
	fecha_atencion = models.DateField(blank=True, null=True)
	hora_atencion = models.TimeField(blank=True, null=True)

	def __unicode__(self):
		return u'%s' % self.codigo

	def toDict(self,recursive = False):
		if recursive:
			return {
				'pk': self.pk,
				'solicitud_inspeccion': self.solicitud_inspeccion.toDict(),
				'codigo': self.codigo,
				'fecha_atencion': datetime.strftime(self.fecha_atencion, '%d-%m-%Y'),
				'hora_atencion': self.hora_atencion.strftime('%I:%M %p'),
			}
		else:
			return {
				'pk': self.pk,
				'solicitud_inspeccion': self.solicitud_inspeccion.pk,
				'codigo': self.codigo,
				'fecha_atencion': datetime.strftime(self.fecha_atencion, '%d-%m-%Y'),
				'hora_atencion': self.hora_atencion.strftime('%I:%M %p'),
			}

	@staticmethod
	def reporte(filtros):
		condiciones = []
		fechaInicioSol = filtros.get('fecha_inicio_sol', None)
		fechaFinSol = filtros.get('fecha_fin_sol', None)
		fechaInicioAten = filtros.get('fecha_inicio_aten', None)
		fechaFinAten = filtros.get('fecha_fin_aten', None)
		estatus = filtros.get('estatus', None)
		estado = filtros.get('estado', None)
		municipio = filtros.get('municipio', None)
		centro = filtros.get('centro', None)

		numeros_orden = NumeroOrden.objects.all()
		if filtros:
			if fechaInicioSol and fechaFinSol:
				fechaInicioSol = datetime.strptime(fechaInicioSol, '%d/%m/%Y')
				fechaFinSol = datetime.strptime(fechaFinSol, '%d/%m/%Y')
				if fechaInicioSol <= fechaFinSol:
					condiciones.append(Q(solicitud_inspeccion__fecha_creacion__range=(fechaInicioSol,fechaFinSol)))

			if fechaInicioAten and fechaFinAten:
				fechaInicioAten = datetime.strptime(fechaInicioAten, '%d/%m/%Y')
				fechaFinAten = datetime.strptime(fechaFinAten, '%d/%m/%Y')
				if fechaInicioAten <= fechaFinAten:
					condiciones.append(Q(fecha_atencion__range=(fechaInicioAten,fechaFinAten)))

			if estatus:
				condiciones.append(Q(solicitud_inspeccion__estatus__id = estatus))

			if estado:
				condiciones.append(Q(solicitud_inspeccion__centro_inspeccion__municipio__estado__id = estado))

			if municipio:
				condiciones.append(Q(solicitud_inspeccion__centro_inspeccion__municipio__id = municipio))

			if centro:
				condiciones.append(Q(solicitud_inspeccion__centro_inspeccion__id = centro))

			numeros_orden = numeros_orden.filter(reduce(operator.and_, condiciones)).order_by('fecha_atencion')
			
		return numeros_orden

	@staticmethod
	def generarReporteXls(numeros_orden):
		"""Método que construye el Excel para la exportación"""
		wb = Workbook()
		ws = wb.active
		ws.title = "Numeros de Orden"
		ws['A1'] = 'Fecha de solicitud'
		#ws['A1'].style.alignment.warp_text = True
		ws['B1'] = 'Fecha de atención'
		ws['C1'] = 'Usuario'
		ws['D1'] = 'Estatus'
		ws['E1'] = 'Centro'
		ws['F1'] = 'Estado'
		ws['G1'] = 'Municipio'

		for index, item in enumerate(numeros_orden):
			ws['A'+str(index+2)] = item.solicitud_inspeccion.fecha_creacion.strftime('%d-%m-%Y')
			ws['B'+str(index+2)] = item.fecha_atencion.strftime('%d-%m-%Y')
			ws['C'+str(index+2)] = item.solicitud_inspeccion.usuario.nombres + ' ' + item.solicitud_inspeccion.usuario.apellidos
			ws['D'+str(index+2)] = item.solicitud_inspeccion.estatus.nombre
			ws['E'+str(index+2)] = item.solicitud_inspeccion.centro_inspeccion.nombre
			ws['F'+str(index+2)] = item.solicitud_inspeccion.centro_inspeccion.municipio.estado.nombre
			ws['G'+str(index+2)] = item.solicitud_inspeccion.centro_inspeccion.municipio.nombre

		# ws['A1'].wrap_text = True

		return wb


class TipoEncuesta(models.Model):
	codigo = models.CharField(max_length=50)
	descripcion = models.CharField(max_length=255)

	def __unicode__(self):
		return u'%s' % self.descripcion


class Encuesta(models.Model):
	nombre = models.CharField(max_length=255)
	preguntas = models.ManyToManyField('Pregunta')
	#usuarios = models.ManyToManyField(USER_MODEL)
	descripcion = models.TextField(blank=True, null=True)
	tipo_encuesta = models.ForeignKey(TipoEncuesta, null=True)

	def __unicode__(self):
		return u'%s' % self.nombre

	@staticmethod
	def estadisticas(filtros):
		"""Método que retorna las estadísticas de las respuestas de las encuestas"""
		matriz = {}
		condiciones = []
		id_encuesta = filtros.get('encuesta', None)

		if id_encuesta:
			condiciones.append(Q(id = id_encuesta))
			encuesta = Encuesta.objects.filter(reduce(operator.and_, condiciones)).first()
			preguntas = Pregunta.objects.filter(encuesta = encuesta, tipo_respuesta__codigo = 'RESP_DEF')
			for p in preguntas:
				matriz[p.enunciado] = {}
				valores_pregunta_encuesta = ValorPreguntaEncuesta.objects.filter(pregunta=p, encuesta = encuesta)
				for vpe in valores_pregunta_encuesta:
					matriz[p.enunciado][vpe.valor] = RespuestaDefinida.objects.filter(respuesta__encuesta = encuesta,valor_definido = vpe.valor, respuesta__pregunta = p).count()

		return matriz

	@staticmethod
	def generarEstadisticasEncuestasXls(matriz):
		"""Método que retorna las encuestas respondidas por los usuarios"""
		wb = Workbook()
		ws = wb.active
		ws.title = "Estadisticas de encuestas"
		i = 1
		j = 'B'
		for key, value in matriz.items():
			j = 'B'
			for k,v in value.items():
				ws[j+str(i)] = k.valor
				j = chr(ord(j)+1)

			j = 'B'
			i+=1
			ws['A'+str(i)] = key
			for k,v in value.items():
				ws[j+str(i)] = v
				j = chr(ord(j)+1)

			i+=1

		return wb



class TipoRespuesta(models.Model):
	codigo = models.CharField(max_length=50)
	descripcion = models.CharField(max_length=255)

	def __unicode__(self):
		return u'%s' % self.descripcion


class Pregunta(models.Model):
	enunciado = models.CharField(max_length=255)
	tipo_respuesta = models.ForeignKey(TipoRespuesta, null=True)

	def __unicode__(self):
		return u'%s' % self.enunciado


class ValorPosible(models.Model):
	valor = models.CharField(max_length=255)
	valor_pregunta_encuesta = models.ManyToManyField('Pregunta', through='ValorPreguntaEncuesta', related_name='valores_pregunta_encuesta')

	def __unicode__(self):
		return u'%s' % self.valor


class ValorPreguntaEncuesta(models.Model):
    valor = models.ForeignKey(ValorPosible, related_name='valor_pregunta')
    pregunta = models.ForeignKey(Pregunta, related_name='pregunta_encuesta')
    encuesta = models.ForeignKey(Encuesta, related_name='encuesta')
    orden = models.IntegerField(default=0)

    def __unicode__(self):
        return u'%s : %s ->  %s' % (self.valor, self.pregunta, self.encuesta)


class Respuesta(models.Model):
	encuesta = models.ForeignKey(Encuesta, default=None)
	pregunta = models.ForeignKey(Pregunta)
	usuario = models.ForeignKey(USER_MODEL, null=True)
	notificacion_usuario = models.ForeignKey('NotificacionUsuario')

	def __unicode__(self):
		return u'%s' % self.id


class RespuestaIndefinida(models.Model):
	respuesta = models.ForeignKey(Respuesta)
	valor_indefinido = models.CharField(max_length=255)

	def __unicode__(self):
		return u'%s' % self.valor_indefinido


class RespuestaDefinida(models.Model):
	respuesta = models.ForeignKey(Respuesta)
	valor_definido = models.ForeignKey(ValorPosible)

	def __unicode__(self):
		return u'%s' % self.valor_definido


# class ColaAtencion(models.Model):
# 	centro_inspeccion = models.ForeignKey(CentroInspeccion)
# 	numero_orden = models.ForeignKey(NumeroOrden)
# 	orden = models.IntegerField()


class Poliza(models.Model):
	cedula_cliente = models.CharField(max_length = 100)
	fecha_inicio_vigencia = models.DateField()
	fecha_fin_vigencia = models.DateField()
	numero = models.IntegerField()
	usuario = models.ForeignKey(USER_MODEL, blank = True, null = True)

	def __unicode__(self):
		return u'%s' % self.numero


class Estatus(models.Model):
	nombre = models.CharField(max_length=255)
	codigo = models.CharField(max_length=100)

	def __unicode__(self):
		return u'%s' % self.nombre


class Notificacion(models.Model):
	asunto = models.CharField(max_length=255, default="Sin Asunto")
	mensaje = models.TextField()
	tipo_notificacion = models.ForeignKey('TipoNotificacion')
	notificacion_usuario = models.ManyToManyField(USER_MODEL, through='NotificacionUsuario', related_name='notificacion_usuario')
	encuesta = models.ForeignKey(Encuesta, blank=True, null=True)

	def __unicode__(self):
		return u'%s' % self.mensaje



class NotificacionUsuario(models.Model):
	notificacion = models.ForeignKey(Notificacion, related_name='notificacion')
	usuario = models.ForeignKey(USER_MODEL, related_name='usuario')
	fecha_creacion = models.DateTimeField(auto_now_add=True)
	leida = models.BooleanField(default=False)
	borrada = models.BooleanField(default=False)
	encuesta_respondida = models.BooleanField(default=False)

	def __unicode__(self):
		return u'%s <- %s' % (self.usuario, self.notificacion)

	@staticmethod
	def encuestas_resueltas(filtros):
		"""Método que retorna las encuestas respondidas por los usuarios"""
		condiciones = []
		usuario_nombres = filtros.get('usuario_nombres', None)
		usuario_apellidos = filtros.get('usuario_apellidos', None)
		tipo_encuesta = filtros.get('tipo_encuesta', None)
		encuesta = filtros.get('encuesta', None)

		if filtros:
			if usuario_nombres:
				condiciones.append(Q(usuario__nombres__icontains = usuario_nombres))

			if usuario_apellidos:
				condiciones.append(Q(usuario__apellidos__icontains = usuario_apellidos))

			if tipo_encuesta:
				condiciones.append(Q(notificacion__encuesta__tipo_encuesta__id = tipo_encuesta))

			if encuesta:
				condiciones.append(Q(notificacion__encuesta__id = encuesta))

		condiciones.append(Q(notificacion__encuesta__isnull = False))
		condiciones.append(Q(encuesta_respondida = True))
		notificaciones = NotificacionUsuario.objects.filter(reduce(operator.and_, condiciones))

		return notificaciones



class TipoNotificacion(models.Model):
	codigo = models.CharField(max_length=100)
	descripcion = models.CharField(max_length=255)

	def __unicode__(self):
		return u'%s' % self.descripcion


class CentrosTiemposAtencion(models.Model):
	fecha = models.DateField()
	centro_inspeccion = models.ForeignKey(CentroInspeccion)
	tiempo_atencion = models.IntegerField()

	def __unicode__(self):
		return u'%s %s %s' % (self.fecha,self.centro_inspeccion,self.tiempo_atencion)


class ParametrosGenerales(models.Model):
	codigo = models.CharField(max_length = 100)
	valor = models.CharField(max_length = 255)

	def __unicode__(self):
		return u'%s:%s' % (self.codigo, self.valor)


class Reclamo(models.Model):
	usuario = models.ForeignKey(USER_MODEL)
	motivo = models.CharField(max_length = 255)
	contenido = models.CharField(max_length = 255)

	def __unicode__(self):
		return u'%s' % self.motivo